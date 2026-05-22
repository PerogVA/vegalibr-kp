# -*- coding: utf-8 -*-
"""Парсер Excel-заявок для ВЕГАЛИБР"""

import re
import openpyxl


def parse_excel(filepath_or_stream):
    """
    Читает Excel-заявку. Возвращает (items, error).
    items = список (name, qty, price_no_kal, kalib)
      price_no_kal — цена из Excel без калибровки (или None)
      kalib        — стоимость калибровки из Excel (или None)
    """
    wb = openpyxl.load_workbook(filepath_or_stream, data_only=True)

    # Выбор листа
    ws = None
    for sname in wb.sheetnames:
        if 'заявк' in sname.lower():
            ws = wb[sname]
            break
    if ws is None:
        ws = wb.active

    # Поиск строки заголовков и колонок
    name_col      = None   # тип (Кольцо резьбовое / Калибр-скоба...)
    spec_col      = None   # обозначение / артикул (содержит М×шаг)
    type_col      = None   # Тип: ПР / НЕ / ПР-НЕ (смета-формат)
    qty_col       = None
    price_col     = None   # цена без калибровки
    price_kal_col = None   # цена с калибровкой
    header_row_idx = None

    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).lower().strip()

            # Наименование/тип (НЕ "с калибровкой")
            if re.search(r'наим', val) and 'калибровк' not in val:
                if name_col is None:
                    name_col = cell.column
                    header_row_idx = cell.row

            # Обозначение / артикул (вторая колонка с описанием)
            if re.search(r'обозн|артикул', val):
                spec_col = cell.column
                header_row_idx = header_row_idx or cell.row

            # Тип (ПР/НЕ — смета-формат)
            if re.fullmatch(r'тип', val):
                type_col = cell.column

            # Кол-во
            if re.search(r'\bкол[\-\.]?во?\b|\bкол\.?\b|\bqty\b|\bcount\b', val) \
                    and '/' not in val and 'цена' not in val and 'руб' not in val:
                qty_col = cell.column

            # Цены — различаем "без калибровки" vs "с калибровкой"
            if ('цен' in val or 'руб' in val) and 'калибровк' in val:
                if 'без калибр' in val:
                    price_col = cell.column      # "цена без калибровки"
                elif 'с калибр' in val:
                    price_kal_col = cell.column  # "цена с калибровкой"

        if name_col and qty_col:
            break

    # Если есть только «Обозначение» без «Наименование» — используем его как name_col
    if name_col is None and spec_col is not None:
        name_col = spec_col
        spec_col = None

    if name_col is None or qty_col is None:
        return [], "Не удалось найти колонки «Наименование» и «Кол-во»"

    # Если spec_col не нашли по заголовку — ищем по содержимому строк
    if spec_col is None and header_row_idx:
        for row in ws.iter_rows(min_row=header_row_idx + 1,
                                max_row=header_row_idx + 15):
            for cell in row:
                if cell.column == name_col or cell.column == qty_col:
                    continue
                if cell.value and re.search(
                        r'[МмMm]\s*\d+[,.]?\d*\s*[×xхХ\-]', str(cell.value)):
                    spec_col = cell.column
                    break
            if spec_col:
                break

    # Читаем строки данных
    items = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        rv = {cell.column: cell.value for cell in row}

        type_name = str(rv.get(name_col) or '').strip()
        spec_name = str(rv.get(spec_col) or '').strip() if spec_col else ''
        side_val  = str(rv.get(type_col) or '').strip().upper() if type_col else ''

        # Предобработка ГОСТ-артикула: «8221-3056-7Н М12×1-7Н» → type_name=«М12×1-7Н», gost_side=«ПР»
        gost_side = ''
        if re.match(r'^[ВBвб]?\d{4}[-‐]\d', type_name):
            m_art = re.match(r'^[ВBвб]?(\d{4})', type_name)
            if m_art:
                _art = m_art.group(1)
                if _art in ('8221', '8311', '8313', '8315'):
                    gost_side = 'ПР'
                elif _art in ('8222', '8312', '8314', '8316'):
                    gost_side = 'НЕ'
            # Вырезаем номинал (М... или Пробка/Кольцо/Ø) после артикула
            m_nom = re.search(
                r'(?:[МмMm]\s*\d|[Пп]робка|[Кк]ольцо|Ø|Ø).+', type_name)
            if m_nom:
                type_name = m_nom.group(0).strip()

        # «Пробка М2-6Н» / «Кольцо резьбовое» → «Калибр-пробка М2-6Н» / «Калибр-кольцо резьбовое»
        if re.match(r'^[Пп]робка\b', type_name):
            type_name = 'Калибр-пробка ' + type_name[6:].strip()
        elif re.match(r'^[Кк]ольцо\b', type_name):
            type_name = 'Калибр-кольцо ' + type_name[6:].strip()

        # Результирующее ПР/НЕ: из колонки Тип (приоритет) или из артикула ГОСТ
        eff_side = side_val or gost_side

        if not type_name and not spec_name:
            continue

        # Пропускаем строки-разделители/заголовки секций (нет цифры в колонке qty)
        raw_qty = rv.get(qty_col)
        try:
            float(str(raw_qty or ''))
        except (ValueError, TypeError):
            # не число → скорее всего заголовок секции, пропускаем
            continue

        # Определяем: является ли name_col просто резьбовым обозначением (смета-формат)?
        # Смета-формат: наименование = «М4×0,7-6g» без указания типа калибра
        name_is_spec = bool(re.match(r'^[МмMm]\s*\d', type_name)) and \
                       not re.search(r'пробка|кольцо|калибр|скоба', type_name, re.I)

        if name_is_spec:
            # Смета-формат: тип калибра нужно вывести из квалитета + тип_col (ПР/НЕ)
            thread_spec = type_name  # например «М4×0,7-6g»
            # Определяем тип по квалитету
            qual_match = re.search(
                r'[МмMm]\s*\d+[,.]?\d*\s*[×xхХ]?\s*\d*[,.]?\d*\s*-\s*([A-Za-zА-Яа-яЁё0-9]+)',
                thread_spec)
            caliber_type = ''
            if qual_match:
                qual = qual_match.group(1)
                if re.search(r'[gefhdr]', qual):   # h — тоже вал (кольцо/контрольник)
                    # КПР/КНЕ/КИ — контрольные калибры-пробки с квалитетом g
                    if re.search(r'\bК[ПН]Р\b|\bКИ\b', thread_spec, re.IGNORECASE):
                        caliber_type = 'Контрольный калибр-пробка'
                    else:
                        caliber_type = 'Калибр-кольцо'
                elif re.search(r'[HН]', qual):
                    caliber_type = 'Калибр-пробка'
            # Если тип не определён по квалитету — пропускаем (нет смысла)
            if not caliber_type:
                continue
            # Добавляем ПР/НЕ
            if eff_side in ('ПР', 'НЕ', 'ПР-НЕ', 'ПР/НЕ'):
                name = f"{caliber_type} {thread_spec} {eff_side}"
            else:
                name = f"{caliber_type} {thread_spec}"

        elif spec_name and re.search(r'[Мм]\s*\d', spec_name):
            # spec уже содержит тип калибра — не дублируем type_name
            spec_has_type = bool(re.search(r'пробка|кольцо|калибр|скоба', spec_name, re.I))
            if spec_has_type or not type_name:
                name = spec_name
            else:
                # берём только тип из type_name (первые 1-2 слова без ГОСТ/артикулов)
                type_clean = re.sub(r'\s*ГОСТ\s*\d+[\.\-–]\d+', '', type_name, flags=re.I)
                type_clean = re.sub(r'\b\d{3,}[\-–]\d{3,}[\-–]?\w*\b', '', type_clean).strip()
                type_clean = type_clean.split()[0] if type_clean else ''  # только первое слово
                name = f"{type_clean} {spec_name}".strip() if type_clean else spec_name
                # Добавляем ПР/НЕ
                if eff_side in ('ПР', 'НЕ', 'ПР-НЕ', 'ПР/НЕ'):
                    name = f"{name} {eff_side}"
        else:
            name = type_name or spec_name
            # Добавляем ПР/НЕ
            if eff_side in ('ПР', 'НЕ', 'ПР-НЕ', 'ПР/НЕ') and \
                    not re.search(r'\b(ПР|НЕ)\b', name.upper()):
                name = f"{name} {eff_side}"

        # Убираем ГОСТ, артикулы и "×Nшт" из имени
        # ("×Nшт" — шаблонное кол-во в ячейке; реальное кол-во берём из колонки Кол-во)
        name = re.sub(r'\s*ГОСТ\s*\d+[\.\-–]\d+', '', name, flags=re.I)
        name = re.sub(r'\b\d{3,}[\-–]\d{3,}[\-–]?\w*\b', '', name)
        name = re.sub(r'[-×х]\s*\d+\s*шт\.?', '', name, flags=re.I)
        name = ' '.join(name.split())
        if not name or name.isdigit():
            continue

        # Количество (raw_qty уже прошёл проверку float выше)
        try:
            qty = int(float(str(raw_qty)))
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        # Цены из Excel
        price_from_excel = None
        kalib_from_excel = None
        if price_col:
            try:
                price_from_excel = float(str(rv.get(price_col) or ''))
            except (ValueError, TypeError):
                pass
        if price_col and price_kal_col:
            try:
                p_no  = float(str(rv.get(price_col)     or ''))
                p_kal = float(str(rv.get(price_kal_col) or ''))
                kalib_from_excel = round(p_kal - p_no, 2)
            except (ValueError, TypeError):
                pass

        items.append((name, qty, price_from_excel, kalib_from_excel))

    if not items:
        return [], "Строки с данными не найдены"

    return items, None
