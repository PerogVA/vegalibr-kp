# -*- coding: utf-8 -*-
"""Парсер Excel-заявок для ВЕГАЛИБР"""

import re
import openpyxl

def parse_zaявka(filepath_or_stream):
    """
    Читает Excel-заявку, возвращает список (name, qty).
    Ищет листы с 'заявк' в названии, иначе берёт первый.
    Определяет колонки по заголовкам.
    """
    wb = openpyxl.load_workbook(filepath_or_stream, data_only=True)

    # Выбор листа
    ws = None
    for sname in wb.sheetnames:
        if 'заявк' in sname.lower():
            ws = wb[sname]
            break
    if ws is None:
        ws = wb.active

    # Поиск строки заголовков и нужных колонок
    name_col = None
    qty_col  = None
    header_row_idx = None

    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).lower().strip()
            # Колонка с наименованием: содержит "наим", но не "калибровк"
            if re.search(r'наим', val) and 'калибровк' not in val:
                name_col = cell.column
                header_row_idx = cell.row
            # Колонка с количеством: "кол-во" / "кол." / "кол" / "qty"
            # НЕ должно содержать "/" (руб/шт) или "цена"
            if re.search(r'\bкол[\-\.]?во?\b|\bqty\b|\bcount\b', val) and '/' not in val and 'цена' not in val:
                qty_col = cell.column
        if name_col and qty_col:
            break

    if name_col is None or qty_col is None:
        return [], "Не удалось найти колонки «Наименование» и «Кол-во» в файле"

    # Читаем строки данных
    items = []
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_vals = {cell.column: cell.value for cell in row}
        name = row_vals.get(name_col)
        qty  = row_vals.get(qty_col)

        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()
        if not name_str or name_str.isdigit():
            continue

        try:
            qty_int = int(float(str(qty)))
        except (ValueError, TypeError):
            continue
        if qty_int <= 0:
            continue

        items.append((name_str, qty_int))

    if not items:
        return [], "Строки с данными не найдены"

    return items, None
